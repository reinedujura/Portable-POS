"""Export functionality for transactions and customers to multiple formats"""

from typing import List, Dict, Optional, Tuple
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

# Try to import optional dependencies
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def generate_sales_excel(transactions: List[Dict], business_name: str) -> Optional[bytes]:
    """
    Generate Excel file from transaction data.
    
    Args:
        transactions: List of transaction dictionaries
        business_name: Name of the business
        
    Returns:
        Excel file as bytes or None if openpyxl not available
    """
    if not EXCEL_AVAILABLE:
        logger.warning("openpyxl not available for Excel export")
        return None
    
    from io import BytesIO
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    
    # Add title
    ws['A1'] = f"Sales Report - {business_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')
    
    # Add headers
    headers = ['Transaction ID', 'Date', 'Customer', 'Items', 'Subtotal', 'Tax', 'Discount', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row, transaction in enumerate(transactions, 4):
        ws.cell(row=row, column=1).value = str(transaction.get('_id', ''))
        ws.cell(row=row, column=2).value = str(transaction.get('date', ''))
        ws.cell(row=row, column=3).value = transaction.get('customer_name', 'Walk-in')
        ws.cell(row=row, column=4).value = len(transaction.get('items', []))
        ws.cell(row=row, column=5).value = float(transaction.get('subtotal', 0))
        ws.cell(row=row, column=6).value = float(transaction.get('tax_amount', 0))
        ws.cell(row=row, column=7).value = float(transaction.get('discount_amount', 0))
        ws.cell(row=row, column=8).value = float(transaction.get('final_total', 0))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def generate_customers_excel(customers: List[Dict], business_name: str) -> Optional[bytes]:
    """
    Generate Excel file from customer data.
    
    Args:
        customers: List of customer dictionaries
        business_name: Name of the business
        
    Returns:
        Excel file as bytes or None if openpyxl not available
    """
    if not EXCEL_AVAILABLE:
        logger.warning("openpyxl not available for Excel export")
        return None
    
    from io import BytesIO
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Add title
    ws['A1'] = f"Customers - {business_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:G1')
    
    # Add headers
    headers = ['Customer ID', 'Name', 'Email', 'Phone', 'Address', 'Total Spent', 'Last Purchase']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row, customer in enumerate(customers, 4):
        ws.cell(row=row, column=1).value = str(customer.get('_id', ''))
        ws.cell(row=row, column=2).value = customer.get('name', '')
        ws.cell(row=row, column=3).value = customer.get('email', '')
        ws.cell(row=row, column=4).value = customer.get('phone', '')
        ws.cell(row=row, column=5).value = customer.get('address', '')
        ws.cell(row=row, column=6).value = float(customer.get('total_spent', 0))
        ws.cell(row=row, column=7).value = str(customer.get('last_purchase', ''))
    
    # Adjust column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def generate_summary_stats(transactions: List[Dict]) -> Dict:
    """
    Generate summary statistics from transactions.
    
    Args:
        transactions: List of transaction dictionaries
        
    Returns:
        Dictionary containing summary statistics
    """
    if not transactions:
        return {
            'total_sales': 0,
            'total_transactions': 0,
            'average_transaction': 0,
            'total_items': 0,
            'total_discount': 0,
            'total_tax': 0
        }
    
    total_sales = sum(float(t.get('final_total', 0)) for t in transactions)
    total_items = sum(len(t.get('items', [])) for t in transactions)
    total_discount = sum(float(t.get('discount_amount', 0)) for t in transactions)
    total_tax = sum(float(t.get('tax_amount', 0)) for t in transactions)
    
    return {
        'total_sales': total_sales,
        'total_transactions': len(transactions),
        'average_transaction': total_sales / len(transactions) if transactions else 0,
        'total_items': total_items,
        'total_discount': total_discount,
        'total_tax': total_tax,
        'items_per_transaction': total_items / len(transactions) if transactions else 0
    }


def calculate_daily_breakdown(transactions: List[Dict]) -> Dict[str, float]:
    """
    Calculate daily sales breakdown.
    
    Args:
        transactions: List of transaction dictionaries
        
    Returns:
        Dictionary with dates as keys and daily totals as values
    """
    daily_sales = {}
    
    for transaction in transactions:
        date = str(transaction.get('date', 'Unknown'))[:10]  # YYYY-MM-DD format
        total = float(transaction.get('final_total', 0))
        
        if date not in daily_sales:
            daily_sales[date] = 0
        daily_sales[date] += total
    
    return daily_sales
